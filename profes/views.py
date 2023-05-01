from django.shortcuts import render, get_object_or_404, redirect
from .models import Materia, Profesor, Tema
from .forms import ReporteForm
from django.db.models import Count, F
import io
import zipfile
from django.http import HttpResponse
from openpyxl import Workbook
from django.core.files.base import ContentFile
def index(request):
    profesores = Profesor.objects.all()
    for profesor in profesores:
        temas_vistos = profesor.temas_vistos.all()
        total_temas = Tema.objects.filter(materia=profesor.materia).count()
        porcentaje_avance = round(temas_vistos.count() / total_temas * 100, 2)
        profesor.porcentaje_avance = porcentaje_avance
    return render(request, 'profes/index.html', {'profesores': profesores})

def editar_reporte(request, matricula, grupo):
    print(matricula)
    try:
        matricula = int(matricula)
    except ValueError:
        return render(request, 'profes/error.html', {'mensaje': 'La matrícula del profesor debe ser un número entero'})

    profesor = Profesor.objects.filter(matricula=matricula, grupo=grupo).first()
    temas = Tema.objects.filter(materia=profesor.materia)
    porcentaje = 0

    if request.method == 'POST':
        form = ReporteForm(request.POST, profesor=profesor)
        if form.is_valid():
            form.save(profesor)
            temas_vistos = profesor.temas_vistos.all()
            total_temas = temas.count()
            porcentaje = temas_vistos.count() / total_temas * 100
        
            
            return redirect('index')
    else:
        form = ReporteForm(profesor=profesor, initial={
            f'tema_{tema.id}': tema.id in profesor.temas_vistos.all().values_list('id', flat=True)
            for tema in temas
        })
        temas_seleccionados = len([k for k, v in request.POST.items() if k.startswith('tema_') and v == 'on'])

    return render(request, 'profes/editar_reporte.html', {'profesor': profesor, 'temas': temas, 'form':form, 'porcentaje_avance': porcentaje})


def detalle_profesor(request, pk):
    profesor = get_object_or_404(Profesor, pk=pk)
    materia = profesor.materia
    temas = Tema.objects.filter(materia=materia)
    context = {
        'profesor': profesor,
        'materia': materia,
        'temas': temas,
    }
    return render(request, 'profes/detalle_profesor.html', context)



def descargar_reportes(request):
    profesores = Profesor.objects.all()
    
    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de avance"
    ws.cell(row=1, column=1, value="Matricula")
    ws.cell(row=1, column=2, value="Nombre")
    ws.cell(row=1, column=3, value="Grupo")
    ws.cell(row=1, column=4, value="Materia")
    ws.cell(row=1, column=5, value="Tema")
    ws.cell(row=1, column=6, value="Visto")
    
    # Agregar los datos de cada profesor al archivo Excel
    row_num = 2
    for profesor in profesores:
        temas_vistos = profesor.temas_vistos.all()
        total_temas = Tema.objects.filter(materia=profesor.materia).count()
        porcentaje = temas_vistos.count() / total_temas * 100
        
        for tema in Tema.objects.filter(materia=profesor.materia):
            tema_visto = 'Sí' if tema in temas_vistos else 'No'
            ws.cell(row=row_num, column=1, value=profesor.matricula)
            ws.cell(row=row_num, column=2, value=profesor.nombre)
            ws.cell(row=row_num, column=3, value=profesor.grupo)
            ws.cell(row=row_num, column=4, value=profesor.materia.nombre)
            ws.cell(row=row_num, column=5, value=tema.nombre)
            ws.cell(row=row_num, column=6, value=tema_visto)
            row_num += 1
    
    # Guardar archivo Excel en memoria
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Crear archivo zip y agregar archivo Excel
    zip_file = io.BytesIO()
    with zipfile.ZipFile(zip_file, mode='w') as zf:
        zf.writestr('Reporte de avance.xlsx', excel_file.getvalue())
    
    # Descargar archivo zip
    response = HttpResponse(zip_file.getvalue(), content_type='application/x-zip-compressed')
    response['Content-Disposition'] = 'attachment; filename=reportes.zip'
    return response